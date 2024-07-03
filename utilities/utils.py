import logging

import softest
from openpyxl import Workbook, load_workbook
import csv


class Utils(softest.TestCase):
    def assertlist_itemtext(self,list,value):
        for stop in list:

            print("The text is: "+stop.text)

            #assert stop.text==value
            self.soft_assert(self.assertEqual,stop.text,value)
            if stop.text==value:
                print("test passed")
            else:
                print("test failed")
            #print("assert pass")
        self.assert_all()
    def sample_logger(loglevel=logging.DEBUG):
        #create logger
        logger=logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)
        #create console handler or file handler and set log level
        ch=logging.StreamHandler()
        fh=logging.FileHandler("frameworklogfile.log")
        #create formatter
        formatter=logging.Formatter('%(asctime)s - %(levelname)s : %(message)s',datefmt='%d/%m/%Y %I:%M:%S %p')
        formatter1=logging.Formatter('%(asctime)s - %(levelname)s : %(message)s')

        #add formatter to console or file handler
        ch.setFormatter(formatter)
        fh.setFormatter(formatter1)
        #add console handler to logger
        logger.addHandler(ch)
        logger.addHandler(fh)
        return logger
    def read_data_from_excel(file_name,sheet):
        datalist=[]
        wb=load_workbook(filename=file_name)
        sh=wb[sheet]
        row_ct=sh.max_row
        col_ct=sh.max_column
        for i in range(2,row_ct+1):
            row=[]
            for j in range(1,col_ct+1):
                row.append(sh.cell(row=i,column=j).value)

            datalist.append(row)
        return datalist
    def read_data_from_csv(filename):
        #create an empty list
        datalist=[]

        #open csv file
        csvdata=open(filename,"r")

        #create csv reader
        reader=csv.reader(csvdata)

        #skip header
        next(reader)

        #add csv rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist


